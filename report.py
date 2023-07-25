from enum import Enum
import datetime
from database import generate_report
from telebot import types
import openpyxl
import io

class Report:
    def __init__(self):
        self.today = datetime.date.today()
        self.start_date = None
        self.end_date = None
        self.report_type = None

    def process_message(self, message, chat_id):
        if message.text == 'Отчеты':
            keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
            button_daily = types.KeyboardButton("Ежедневный")
            button_weekly = types.KeyboardButton("Еженедельный")
            button_monthly = types.KeyboardButton("Ежемесячный")
            button_custom = types.KeyboardButton("Другой")
            button_cancel = types.KeyboardButton("Отмена")
            keyboard.add(button_daily, button_weekly, button_monthly, button_custom, button_cancel)
            self.report_type = 'start'
            return None, "Выберите тип отчета:", keyboard
        elif self.report_type == 'start':
            if message.text == 'Отмена':
                self.report_type = None
                return "Отменено."
            elif message.text in ["Ежедневный", "Еженедельный", "Ежемесячный"]:
                self.report_type = message.text
                start_date, end_date = self.get_report_dates(self.report_type)
                report_data = generate_report(start_date, end_date)
                wb = self.generate_excel_report(report_data)
                response = "Отчет готов."
                self.reset()
                return wb, response, None
            elif message.text == "Другой":
                self.report_type = 'custom'
                return "Введите дату начала отчета в формате ГГГГ-ММ-ДД:"
            else:
                return "Пожалуйста, выберите тип отчета:"
        elif self.report_type == 'custom':
            try:
                self.start_date = datetime.datetime.strptime(message.text, "%Y-%m-%d").date()
                self.report_type = 'custom_end'
                return "Введите дату окончания отчета в формате ГГГГ-ММ-ДД"
            except ValueError:
                return "Неверный формат даты. Пожалуйста, введите дату в формате ГГГГ-ММ-ДД"
        elif self.report_type == 'custom_end':
            try:
                self.end_date = datetime.datetime.strptime(message.text, "%Y-%m-%D").date()
                report_data = generate_report(self.start_date, self.end_date)
                wb = self.generate_excel_report(report_data)
                response = "Отчет готов."
                self.reset()
                return wb, response, None
            except ValueError:
                return "Неверный формат даты. Пожалуйста, введите дату в формате ГГГГ-ММ-ДД"
        else:
            None, "Неожиданное состояние", None

    def reset(self):
        self.start_date = None
        self.end_date = None
        self.report_type = None

    def get_report_dates(self, report_type):
        if report_type == "Ежедневный":
            start_date = self.today - datetime.timedelta(days=1)
            end_date = self.today
        elif report_type == "Еженедельный":
            start_date = self.today - datetime.timedelta(days=7)
            end_date = self.today
        elif report_type == "Ежемесячный":
            start_date = self.today - datetime.timedelta(days=30)
            end_date = self.today
        elif report_type == "Другой":
            start_date = self.start_date
            end_date = self.end_date
        return start_date, end_date

    def generate_excel_report(self, report_data):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet['A1'] = 'Название товара'
        sheet['B1'] = 'Продавец'
        sheet['C1'] = 'Магазин'
        sheet['D1'] = 'Сумма продаж'
        sheet['E1'] = 'Дата продажи'
        sheet['F1'] = 'Тип расчета'
        for i, row in enumerate(report_data):
            sheet.cell(row=i+2, column=1).value = row[0]
            sheet.cell(row=i+2, column=2).value = row[1]
            sheet.cell(row=i+2, column=3).value = row[2]
            sheet.cell(row=i+2, column=4).value = row[3]
            sheet.cell(row=i+2, column=5).value = row[4].strftime('%d.%m.%Y')
            sheet.cell(row=i+2, column=6).value = row[5]
        report_buffer = io.BytesIO()
        wb.save(report_buffer)
        report_buffer.seek(0)  # Move the buffer's pointer to the beginning of the file
        return report_buffer
